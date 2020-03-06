# ------ page. 32 ------
# --- 模組 my_cart
'''
class My_cart:
    def __init__(self):
        self.shopping_list = {}

    def add(self,name='', price=0, num=0):
        if len(name.strip(()) > 0 and num > 0):
            self.shopping_list[name] = [price,num]
            print('add item ', name, ' $', price, ' nume=', num)

    def drop(self,name=''):
        if len(name.strip(()) > 0 and num > 0):
            del shopping_list[name]
            print('drip item ', name)

    def modify(self, name='', price=0, num=0):
        if len(name.strip(()) > 0 and num > 0):
            self.shopping_list[name] = [price, num]
            print('modify item ', name, ' $', price, ' nume=', num)

    def account(self):
        total = 0
        for name in self.shopping_list.keys():
            total += self.shopping_list[name][0]*self.shopping_list[name][1]
        print(total)
        return total

    def list(self):
        print((self.shopping_list))

from my_cart import My_Cart

def main():
    c = My_Cart()
    c.add('apple', 30, 1)
    c.add('juice', 10, 12)
    c.modify('juice', 22, 10)
    c.add('cookie', 40, 2)
    c.drop('apple')
    c.list
    c.account()


if __name__ == "__main__":
    main()

'''
'''
# ------ page. 34 ------
import os
with open('test.txt', 'wt') as f:
    print('for test', file=f)

print('這個檔案存在嗎 ?', os.path.exists('test.txt'))

print('這是個檔案嗎 ?', os.path.isfile('test.txt'))

print('這是個目錄嗎 ?', os.path.isdir('.'))

print('目前的工作目錄', os.getcwd())

print('是絕對路徑嗎 ?', os.path.isabs('C:\\Users\\Big data\\PycharmProjects\\untitled1'))

print('列出某個目錄', os.listdir('C:\\Users\\Big data\\PycharmProjects\\untitled1\\my_cart'))

print()
# ------ page. 36 ------
import os, shutil, glob
if os.path.exists('test_dir'):
    shutil.rmtree('test_dir')

os.mkdir('test_dir')

os.chdir('test_dir')

print(os.path.abspath('test_dir'))

shutil.copy('../test.txt', 'test_copy.txt')

os.listdir('.')

print(glob.glob('te*'))

print()
# --- math模組
import math

num = 1.2345
print('捨去小數', math.trunc(num))

num = 9
print('開根號', math.sqrt(num))

print('(x平方 + y平方)然後開根號', math.hypot(3, 4))

print(math.fabs(-10))

print()
# ------ page. 39 ------
# --- random 模組
import random
print('機率模型種子7', random.seed(500))

print('機率', random.random())

print('1~10隨機整數', random.randrange(10))

print('隨機偶數', random.randrange(0, 101, 2))

print('隨機挑一個', random.choice(['剪刀', '石頭', '布']))

print('權重', (random.choices(['人頭', '字'], cum_weights=(0.8, 1.0), k=10)))

num = ['A', 'J', 'Q', 'K', '10']
random.shuffle(num) ; print('洗牌', num)

print('抽樣三個', random.sample([10, 20, 30, 40, 50], k=3))

print()
# ------ page. 40 ------
# --- time 模組
import time

print(time.time())

print(time.localtime())

print(time.asctime(time.localtime(time.time())))

print(time.asctime(time.gmtime(time.time())))

print('格式化成', time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))

print('格式化成', time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))

print()'''
# ------ page. 41 ------
# --- csv 模組
import csv
food_cal_list = [
    ['種類', '單位', '重量(g)', '熱量(卡)'],
    ['白米飯', '1碗', '205', '225'],
    ['意大利肉醬麵', '1份', '248', '330'],
    ['白吐司', '1片', '25', '75'],
    ['全麥吐司', '1片', '25', '65'],
    ['花生醬', '1湯匙', '16', '95']
]
print(food_cal_list)

with open('../food_cal.csv', mode='w', newline='', encoding='utf-8') as csv_file:
    writer = csv.writer(csv_file)
    writer.writerows(food_cal_list)
    writer.writerow(['果醬', '1湯匙', 18, 50])
    print('Write csv OK!')

with open('../food_cal.csv', mode='r', newline='', encoding='utf-8') as csv_file:
    rows = csv.reader(csv_file, delimiter=',')
    read_csv_list = [row for row in rows]
    print(read_csv_list)

print()
# --- 寫入字典資料到 csv 檔案
import csv

# rowdicts
dict_store_info = [
    {'門市': '民復門市', '電話': '02 2514 0673', '地址': '105台北市松山區民生東路四段100號'},
    {'門市': '民富門市', '電話': '02 2762 2884', '地址': '106台北市松山區民生東路四段200號'},
    {'門市': '明復門市', '電話': '02 2719 6327', '地址': '107台北市松山區民生東路四段300號'}
]
dictrow = {'門市': '冥復門市', '電話': '02 2762 2908', '地址': '108台北市松山區民生東路四段400號'}

# 資料欄位
field_list = list(dict_store_info[0].keys())

with open('../dict_out.csv', 'wt', newline='', encoding='Big5') as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=field_list)
    # 寫入欄位
    writer.writeheader()
    # 寫入多列
    writer.writerows(dict_store_info)
    # 寫入一列
    writer.writerow(dictrow)

print()
# ------ page. 42 ------
# --- 從 csv 檔案讀取資料到字典
dict_read_info = []
with open('../dict_out.csv', 'r', newline='', encoding='Big5')as csvfile:
    rows = csv.DictReader(csvfile)
    for row in rows:
        field_list = list(row.keys())
        dict_read = {}

        for field in field_list:
            dict_read[field] = row[field]
        dict_read_info.append(dict_read)

for item in dict_read_info:
    print(item)

print()
# ------ page. 43 ------
# --- pickle 模組
import pickle
data = [12345, 'Hello', '你好', {'臉書': 'Facebook', 'Goole': '谷哥'}]

bytes_str = pickle.dumps(data)
print(bytes_str)

data_load = pickle.loads(bytes_str)
print(data_load)

print()
# --- 讀檔
import pickle
data = [12345, 'Hello', '你好', {'臉書': 'Facebook', 'Goole': '谷哥'}]

with open('../out.pk', 'wb') as f:
    pickle.dump(data, f)

with open('../out.pk', 'rb') as f:
    load_data = pickle.load(f)

print(load_data)

print()
# ------ page. 45 ------
# --- json 模組操作
import json
dict_data = {
    '產品編號': 'PN100023',
    '產品名稱': 'IPhoneX 掀蓋手機殼',
    '產品規格': ['質感黑', '海軍藍', '櫻花粉'],
    '產品價格': 599
}
print(dict_data)
# --- 轉成JSON字串
json_str = json.dumps(dict_data)

# --- unicode字串
print(json_str)

# --- 把unicode字串轉回
print(json_str.encode('utf-8').decode('unicode_escape'))

# --- 轉JSON回字典
data_load = json.loads(json_str)
print(dict(data_load))

# --- Writing JSON data
with open('../data.json', 'w', encoding='utf-8') as f:
    json.dump(dict_data, f)

# --- Reading data back
with open('../data.json', 'r', encoding='utf-8') as f:
    data_load = json.load(f)
    print(dict(data_load))

print()
'''# --- urllib 模組
from urllib import request, parse

# Base URL being accessed
url = 'http://httpbin.org/get'

# Dictionary of query parameters (if any)
parms = {
    'name1': 'value1',
    'name2': 'value1'
}

# Encode the query string
querystring = parse.urlencode(parms)

# Make a GET request and read the response
u = request.urlopen(url+'?' + querystring)
resp = u.read()

print()'''
# ------ page. 46 ------
# --- 使用 urllib 取得開放資料
from json import loads
from urllib import request, parse

# Base URL being accessed
url = 'https://opendata.cwb.gov.tw/api/v1/rest/datastore/F-D0047-005'

# Dictionary of query parameters (if any)
# Authorization = CWB-F6F2390F-673B-41C2-AB26-98A36BCB912A
parms = {
    'Authorization': 'CWB-F6F2390F-673B-41C2-AB26-98A36BCB912A',
    'formate': 'JSON'
}

# Encode the query string
querystring = parse.urlencode(parms)
print(querystring)
print(url+'?' + querystring)

# Make a GET request and read the response
result = request.urlopen(url+'?' + querystring)
resp = dict(loads(result.read()))
print(resp['records'])

print()
# ------ page. 51 ------
# --- 寫入 Excel 檔案
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment

wb = Workbook()

ws = wb.active
ws.title = 'Sheet1'

wb.create_sheet('Sheet2', index=1)

ws['A1'] = '哈囉'
ws.append(['A', 'B', 'C'])
ws.append([200, 300, 500, '=AVERAGE(A3:C3)'])
ws.cell(row=4, column=1).value = 600
ws.cell(4, 2).value = 700
ws = wb['Sheet2']
ws['A1'] = '測試'
myfont = Font(name='微軟正黑體', size=18, italic=True, color=colors.BLUE, bold=True)
ws['A1'].font = myfont

# 存檔
wb.save('out.xlsx')
print('OK!')

print()
# ------ page. 52 ------
# ---讀取 Excel檔案
from openpyxl import load_workbook

wb = load_workbook('../out.xlsx')

print(wb.sheetnames)
print(wb.active)

ws = wb['Sheet1']

print(ws['A3'].value)
print(ws.cell(row=3,column=1).value)
print(ws.cell(3,1).value)

for row in ws.rows:
    for cell in row:
        print(cell.value)

ws = wb['Sheet2']
print(ws['A1'].value)

print()























